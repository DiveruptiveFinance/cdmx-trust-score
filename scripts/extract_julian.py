#!/usr/bin/env python3
"""Extrae las hojas del XLSX oficial de Julián a los CSVs que consume el front.

Fuente: CDMX_TrustScore_Database (1).xlsx — entregado por Julián el 2026-04-18.
Autoridad: Cuenta Pública CDMX 2018–2024 · Secretaría de Administración y Finanzas.

Salidas:
  - public/data/ejecucion.csv            (alcaldía × año · aprobado + ejercido MDP)
  - public/data/scores-by-year.csv       (alcaldía × año · 4 componentes + total)
  - public/data/scores.csv               (alcaldía · promedio del sexenio)
  - public/data/hallazgos.csv            (alcaldía · sexenio · tipo · narrativa)
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "CDMX_TrustScore_Database (1).xlsx"
OUT = ROOT / "public" / "data"

YEARS = [2018, 2019, 2020, 2021, 2022, 2023, 2024]

ALCALDIAS = [
    "Álvaro Obregón", "Azcapotzalco", "Benito Juárez", "Coyoacán",
    "Cuajimalpa de Morelos", "Cuauhtémoc", "Gustavo A. Madero",
    "Iztacalco", "Iztapalapa", "La Magdalena Contreras",
    "Miguel Hidalgo", "Milpa Alta", "Tláhuac", "Tlalpan",
    "Venustiano Carranza", "Xochimilco",
]


def num(v):
    if v is None or v == "N.A.":
        return None
    if isinstance(v, str):
        v = v.replace(",", "").replace("%", "").strip()
        if not v:
            return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def extract_ejecucion(wb):
    """Hoja '📊 Datos Históricos' → ejecucion.csv (alcaldía × año)."""
    ws = wb["📊 Datos Históricos"]
    # Row 3 tiene los años (col C,E,G,... → 2018,2019,...)
    # Row 4 tiene "Aprobado" / "Ejercido" alternados desde col C
    # Filas 5+ : alcaldías en col B, montos desde col C
    rows = []
    for r in range(5, ws.max_row + 1):
        alc = ws.cell(row=r, column=2).value
        if not alc or alc not in ALCALDIAS:
            continue
        for i, year in enumerate(YEARS):
            col_aprob = 3 + i * 2      # C, E, G, ...
            col_ejerc = col_aprob + 1  # D, F, H, ...
            aprob = num(ws.cell(row=r, column=col_aprob).value)
            ejerc = num(ws.cell(row=r, column=col_ejerc).value)
            rows.append({
                "alcaldia": alc,
                "anio": year,
                "aprobado": round(aprob) if aprob is not None else "",
                "modificado": "",
                "ejercido": round(ejerc) if ejerc is not None else "",
            })
    return rows


def extract_scores_by_year(wb):
    """Hoja '🏆 Scores' → scores-by-year.csv (alcaldía × año con 4 componentes)."""
    ws = wb["🏆 Scores"]
    rows = []
    for r in range(4, ws.max_row + 1):
        alc = ws.cell(row=r, column=2).value
        anio = ws.cell(row=r, column=3).value
        if alc not in ALCALDIAS or anio is None:
            continue
        rows.append({
            "alcaldia": alc,
            "anio": int(anio),
            "score_total": num(ws.cell(row=r, column=4).value),
            "score_presupuesto": num(ws.cell(row=r, column=5).value),
            "score_ministraciones": num(ws.cell(row=r, column=6).value),
            "score_deuda": num(ws.cell(row=r, column=7).value),
            "score_patrimonio": num(ws.cell(row=r, column=8).value),
            "data_faltante": "true" if str(ws.cell(row=r, column=9).value).lower() == "true" else "false",
        })
    return rows


def aggregate_sexenio(year_rows):
    """Promedia el sexenio 2018–2024 por alcaldía (ignorando N.A.)."""
    by_alc: dict[str, list[dict]] = {}
    for row in year_rows:
        by_alc.setdefault(row["alcaldia"], []).append(row)

    out = []
    for alc, subs in by_alc.items():
        def avg(key):
            vals = [s[key] for s in subs if s[key] is not None]
            return round(sum(vals) / len(vals), 1) if vals else None

        total = avg("score_total")
        pres = avg("score_presupuesto")
        deuda = avg("score_deuda")
        # plan = ministraciones en el modelo de Julián (homologo nombre a lo que el front espera)
        plan = avg("score_ministraciones")
        patr = avg("score_patrimonio")
        out.append({
            "alcaldia": alc,
            "sexenio": "2018-2024",
            "score_total": round(total) if total is not None else "",
            "score_presupuesto": pres if pres is not None else "",
            "score_plan": plan if plan is not None else "",
            "score_deuda": deuda if deuda is not None else "",
            "score_patrimonio": patr if patr is not None else "",
            "data_faltante": "true",
        })
    out.sort(key=lambda r: r["alcaldia"])
    return out


def extract_hallazgos(wb, year_rows):
    """Hoja '📝 Hallazgos' → hallazgos.csv con tipo derivado de la tasa de ejecución del año."""
    ws = wb["📝 Hallazgos"]
    rows = []
    for r in range(3, ws.max_row + 1):
        alc = ws.cell(row=r, column=2).value
        anio = ws.cell(row=r, column=3).value
        narr = ws.cell(row=r, column=4).value
        if alc not in ALCALDIAS or not narr:
            continue
        # tipo derivado del score_total del año
        match = next(
            (s for s in year_rows if s["alcaldia"] == alc and s["anio"] == int(anio)),
            None,
        )
        score = match["score_total"] if match else None
        tipo = "logro" if (score is not None and score >= 67) else "pendiente"
        rows.append({
            "alcaldia": alc,
            "sexenio": "2018-2024",
            "tipo": tipo,
            "hallazgo_narrativo": narr.strip(),
        })
    return rows


def write_csv(path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main():
    print(f"→ Abriendo {XLSX.name}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    ejec = extract_ejecucion(wb)
    year_rows = extract_scores_by_year(wb)
    scores = aggregate_sexenio(year_rows)
    hall = extract_hallazgos(wb, year_rows)

    OUT.mkdir(parents=True, exist_ok=True)
    write_csv(OUT / "ejecucion.csv", ejec,
              ["alcaldia", "anio", "aprobado", "modificado", "ejercido"])
    write_csv(OUT / "scores-by-year.csv", year_rows,
              ["alcaldia", "anio", "score_total", "score_presupuesto",
               "score_ministraciones", "score_deuda", "score_patrimonio",
               "data_faltante"])
    write_csv(OUT / "scores.csv", scores,
              ["alcaldia", "sexenio", "score_total", "score_presupuesto",
               "score_plan", "score_deuda", "score_patrimonio", "data_faltante"])
    write_csv(OUT / "hallazgos.csv", hall,
              ["alcaldia", "sexenio", "tipo", "hallazgo_narrativo"])

    print(f"✓ ejecucion.csv        ({len(ejec)} filas)")
    print(f"✓ scores-by-year.csv   ({len(year_rows)} filas)")
    print(f"✓ scores.csv           ({len(scores)} filas)")
    print(f"✓ hallazgos.csv        ({len(hall)} filas)")

    print()
    print("Scores por alcaldía (promedio sexenio):")
    for s in scores:
        total = s["score_total"] if s["score_total"] != "" else "—"
        print(f"  {s['alcaldia']:<25} total={total}  "
              f"pres={s['score_presupuesto']}  min={s['score_plan']}  "
              f"deuda={s['score_deuda']}  patr={s['score_patrimonio']}")


if __name__ == "__main__":
    main()
